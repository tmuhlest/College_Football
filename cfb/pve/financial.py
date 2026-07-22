"""Phase 4: Win-value curve and marginal $/win computation (D10, D11).

The win-value curve is:
  - Discrete over regular-season wins 0–12
  - User-editable in the Excel input workbook
  - Excludes the CFP (by design — keeps cross-program comparisons apples-to-apples)
  - No interpolation: discrete lookup only

marginal_win_value(curve, current_wins) returns the $/win for going from
current_wins to current_wins+1 — used as the exchange rate in the optimizer.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# Approximate conversion: 130 cumulative expected points (EPA) per season ≈ 1 marginal win.
# A player with 130 cumulative points added contributed roughly one additional win
# over an average player. Used to convert player points added (EPA) to wins — i.e.
# WAA (Wins Above Average) = points_added / EPA_PER_WIN — before the $/win curve.
# Based on CFB analytics norms; hardcoded for V1.
EPA_PER_WIN: float = 130.0


def load_win_value_curve(workbook_path: str | Path) -> dict[int, float]:
    """Read the win-value curve from the 'Win Value Curve' tab of the input workbook.

    Expects rows: Column A = wins (0–12), Column B = $/win value for that increment.
    Returns a dict {wins_gained: dollar_value}.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Win Value Curve" not in wb.sheetnames:
        raise ValueError("Input workbook must have a 'Win Value Curve' tab.")

    ws = wb["Win Value Curve"]
    curve: dict[int, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        wins, value = row[0], row[1]
        if wins is None:
            continue
        curve[int(wins)] = float(value or 0.0)

    if not curve:
        raise ValueError("Win Value Curve tab is empty or has no data rows.")

    return curve


def marginal_win_value(curve: dict[int, float], current_wins: int) -> float:
    """Return the $/win for going from current_wins to current_wins+1.

    Uses a discrete lookup with no interpolation (D11).
    If current_wins >= 12 (curve max), returns the value at 12.
    """
    lookup = min(current_wins + 1, 12)
    value = curve.get(lookup)
    if value is None:
        print(f"  WARNING: win-value curve has no entry for win {lookup}. "
              f"Check the Win Value Curve tab — returning $0 for this win level.")
        return 0.0
    return float(value)


def load_preseason_wins(workbook_path: str | Path) -> int:
    """Read the team's preseason win projection from the 'Config' tab."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Config"]
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Preseason Win Projection":
            return int(row[1])
    raise ValueError("'Preseason Win Projection' not found in Config tab.")

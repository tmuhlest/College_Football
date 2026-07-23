"""Phase 6: Excel input/output (FAST-standard 8-tab output report).

Input workbook tabs (team fills in):
  Config          — budget ceiling, preseason win projection
  Win Value Curve — wins 0-12 -> $/win (user-editable, discrete, no CFP)
  Two-Deep        — position, starter_name, backup_name, starter_id, backup_id
  Candidates      — player_id, player, position, current_team, asking_price, user_manual_points

Output workbook tabs (left to right, FAST standard):
  Dashboard          — executive summary, master error flag, budget summary
  Read-Me            — instructions, color key, scope notes
  Inputs             — ONLY place to edit: budget, win curve, EPA/win, asking prices
  Calculations       — auditable step-by-step math per candidate
  Candidate Board    — full candidate output
  Recommended Roster — budget-constrained signing recommendation
  Depth Comparison   — current roster vs. candidates by position
  Raw Data           — full projected points added for all FBS players (reference)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pve.financial import EPA_PER_WIN
from pve.ingest import ELIGIBILITY_WARN_CLASSES


# ---------------------------------------------------------------------------
# Input readers
# ---------------------------------------------------------------------------

def read_input_workbook(path: str | Path) -> dict:
    """Read all input tabs from the team's workbook.

    Returns a dict with keys: budget, preseason_wins, win_curve, depth_chart, candidates.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # Config tab
    ws_cfg = wb["Config"]
    cfg = {r[0]: r[1] for r in ws_cfg.iter_rows(min_row=2, values_only=True) if r[0]}
    # Accept both the new label and the legacy "Budget Ceiling" key for backward compatibility.
    budget = float(cfg.get("Skill Position Portal Budget", cfg.get("Budget Ceiling", 0)))
    preseason_wins = int(cfg.get("Preseason Win Projection", 6))

    # Win Value Curve tab
    # Only accept rows where column A is a number — skips any title/instruction rows
    # that appear in the blank template above the data.
    ws_curve = wb["Win Value Curve"]
    win_curve = {}
    for row in ws_curve.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], (int, float)) and row[0] is not None:
            win_curve[int(row[0])] = float(row[1] or 0.0)

    # Two-Deep tab
    # Filter to valid skill positions to skip title/instruction/header rows from the template.
    _VALID_POSITIONS = {"QB", "RB", "WR", "TE"}
    ws_td = wb["Two-Deep"]
    # Slice to first 5 columns — template has a trailing Notes column that is not consumed.
    depth_rows = [r[:5] for r in ws_td.iter_rows(min_row=2, values_only=True)]
    depth_chart = pd.DataFrame(
        depth_rows,
        columns=["position", "starter_name", "backup_name", "starter_id", "backup_id"],
    )
    depth_chart = depth_chart[depth_chart["position"].isin(_VALID_POSITIONS)].reset_index(drop=True)

    # Candidates tab
    # A valid row passes through if it has EITHER:
    #   (a) a non-empty, non-header player_id (model looks it up in projected_points.parquet), OR
    #   (b) a player name + manual_points (FCS/unmodeled players evaluated via user-asserted points)
    # Phantom template rows (instruction text, column headers, blank rows) are dropped.
    # Slice to first 6 columns — template has a trailing Notes column that is not consumed.
    _NON_DATA_LABELS = {"player_id", "Player ID", "Instructions", ""}
    ws_cand = wb["Candidates"]
    cand_rows = [r[:6] for r in ws_cand.iter_rows(min_row=2, values_only=True)]
    candidates = pd.DataFrame(
        cand_rows,
        columns=["player_id", "player", "position", "current_team", "asking_price", "user_manual_points"],
    )

    def _norm_pid(val):
        """Normalize player_id to a clean string.

        openpyxl returns numeric cells as Python int; when pandas mixes int and
        None in a column it upcasts to float64, turning 5078810 into 5078810.0
        and making str() give '5078810.0' which doesn't match the parquet index.
        Converting early to a canonical string avoids that drift everywhere downstream.
        """
        if val is None:
            return None
        try:
            f = float(val)
            if f == int(f):
                return str(int(f))
        except (TypeError, ValueError):
            pass
        return str(val).strip()

    def _valid_pid(val) -> bool:
        return val is not None and str(val).strip() not in _NON_DATA_LABELS

    def _is_numeric(val) -> bool:
        try:
            float(val)
            return True
        except (TypeError, ValueError):
            return False

    candidates["player_id"] = candidates["player_id"].apply(_norm_pid)

    has_name = (
        candidates["player"].notna()
        & (candidates["player"].astype(str).str.strip() != "")
    )
    has_valid_pid = has_name & candidates["player_id"].apply(_valid_pid)
    has_manual_points = has_name & candidates["user_manual_points"].apply(_is_numeric)
    candidates = candidates[has_valid_pid | has_manual_points].reset_index(drop=True)
    candidates["asking_price"] = pd.to_numeric(candidates["asking_price"], errors="coerce").fillna(0)

    return {
        "budget": budget,
        "preseason_wins": preseason_wins,
        "win_curve": win_curve,
        "depth_chart": depth_chart,
        "candidates": candidates,
    }


# ---------------------------------------------------------------------------
# FAST color constants
#   Blue  = hardcoded input (the only cells a user should type in)
#   Green = cross-sheet formula (references another tab)
#   Black = same-sheet formula
#   Red   = error or eligibility warning
# ---------------------------------------------------------------------------
_BLUE  = Font(color="2F5496")
_GREEN = Font(color="375623")
_BLACK = Font(color="000000")
_RED   = Font(color="C00000", bold=True)

_FILL_DARK_BLUE  = PatternFill("solid", fgColor="1F3864")
_FILL_LIGHT_BLUE = PatternFill("solid", fgColor="DCE6F1")

_FMT_DOLLAR = '$#,##0'
_FMT_WINS   = '+0.000;-0.000;"-"'
_FMT_WINS2  = '+0.00;-0.00;"-"'
_FMT_EPA    = '+0.0;-0.0;"-"'

# Stable Inputs tab cell positions — Calculations formulas hard-code these addresses.
# Do NOT move them without updating every formula in _add_calculations_tab.
_BUDGET_REF    = "'Inputs'!$B$3"
_PRESEASON_REF = "'Inputs'!$B$4"
_EPA_WIN_REF   = "'Inputs'!$B$9"
_WIN_CURVE_REF = "'Inputs'!$A$16:$B$27"
_ASK_REF       = "'Inputs'!$A$42:$B$100"


# ---------------------------------------------------------------------------
# Shared style helpers
# ---------------------------------------------------------------------------

def _hdr(ws, row: int, cols: int, fg: str = "1F3864") -> None:
    fill = PatternFill("solid", fgColor=fg)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _title_row(ws, row: int, text: str, n_cols: int, height: int = 22) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = _FILL_DARK_BLUE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = height
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _section_label(ws, row: int, text: str, n_cols: int) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="1F3864")
    cell.fill = _FILL_LIGHT_BLUE
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _safe_na(val):
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


# ---------------------------------------------------------------------------
# Tab 1: Dashboard
# ---------------------------------------------------------------------------

def _add_dashboard_tab(wb) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "4472C4"

    _title_row(ws, 1, "PLAYER VALUE ENGINE — EXECUTIVE DASHBOARD", 2, height=28)

    ws.cell(row=3, column=1, value="MASTER STATUS").font = Font(bold=True)
    master_f = (
        '=IF(AND(B15="OK",B16="OK",B17="OK",B18="OK"),'
        '"✓  MODEL OK","⚠  REVIEW ERROR CHECKS BELOW")'
    )
    ws.cell(row=3, column=2, value=master_f).font = Font(bold=True)
    ws.conditional_formatting.add(
        "B3",
        FormulaRule(
            formula=['B3="✓  MODEL OK"'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
            font=Font(bold=True, color="375623"),
        ),
    )
    ws.conditional_formatting.add(
        "B3",
        FormulaRule(
            formula=['LEFT(B3,1)="⚠"'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(bold=True, color="C00000"),
        ),
    )

    _section_label(ws, 5, "BUDGET SUMMARY", 2)

    ws.cell(row=6, column=1, value="Total Budget").font = Font(bold=True)
    c = ws.cell(row=6, column=2, value=f"={_BUDGET_REF}")
    c.font = _GREEN
    c.number_format = _FMT_DOLLAR

    ws.cell(row=7, column=1, value="Recommended Spend").font = Font(bold=True)
    c = ws.cell(row=7, column=2,
        value="=SUMPRODUCT(('Recommended Roster'!G4:G100=\"Yes\")*'Recommended Roster'!D4:D100)")
    c.font = _GREEN
    c.number_format = _FMT_DOLLAR

    ws.cell(row=8, column=1, value="Budget Remaining").font = Font(bold=True)
    c = ws.cell(row=8, column=2, value="=B6-B7")
    c.font = _BLACK
    c.number_format = _FMT_DOLLAR

    _section_label(ws, 10, "ROSTER SUMMARY", 2)

    ws.cell(row=11, column=1, value="Recommended Signings").font = Font(bold=True)
    c = ws.cell(row=11, column=2, value='=COUNTIF(\'Recommended Roster\'!G4:G100,"Yes")')
    c.font = _GREEN

    ws.cell(row=12, column=1, value="Total Marginal Wins (Recommended)").font = Font(bold=True)
    c = ws.cell(row=12, column=2,
        value='=SUMPRODUCT((\'Recommended Roster\'!G4:G100="Yes")*IFERROR(\'Recommended Roster\'!C4:C100,0))')
    c.font = _GREEN
    c.number_format = _FMT_WINS2

    _section_label(ws, 14, "ERROR CHECKS", 2)

    checks = [
        (15, "Budget not exceeded",
         '=IF(B7<=B6,"OK","⚠ Recommended spend exceeds budget")'),
        (16, "Required inputs present",
         '=IF(AND(ISNUMBER(\'Inputs\'!$B$3),\'Inputs\'!$B$3>0,ISNUMBER(\'Inputs\'!$B$4)),'
         '"OK","⚠ Budget or Preseason Wins missing in Inputs tab")'),
        (17, "Candidate data quality",
         '=IF(COUNTIF(\'Candidate Board\'!E2:E100,"insufficient")=0,"OK",'
         'COUNTIF(\'Candidate Board\'!E2:E100,"insufficient")&'
         '" candidate(s) lack model data — manual points estimate required. See Data Source column.")'),
        (18, "Calculation integrity",
         '=IF(SUMPRODUCT((ISERROR(\'Calculations\'!F3:F100))*(\'Calculations\'!A3:A100<>""))=0,'
         '"OK","⚠ Errors in Calculations tab — check player IDs")'),
    ]

    gf = PatternFill("solid", fgColor="C6EFCE")
    rf = PatternFill("solid", fgColor="FFC7CE")
    for r, label, formula in checks:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=formula).font = _GREEN
        ws.conditional_formatting.add(
            f"B{r}",
            FormulaRule(formula=[f'B{r}="OK"'], fill=gf, font=Font(bold=True, color="375623")),
        )
        ws.conditional_formatting.add(
            f"B{r}",
            FormulaRule(formula=[f'B{r}<>"OK"'], fill=rf, font=Font(bold=True, color="C00000")),
        )

    note = ws.cell(row=20, column=1,
        value="Lambda (risk aversion) is set at the Python level. "
              "Update the reference value in Inputs if you change it, then re-run the pipeline.")
    note.font = Font(italic=True, color="595959")
    ws.merge_cells("A20:B20")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45


# ---------------------------------------------------------------------------
# Tab 2: Read-Me
# ---------------------------------------------------------------------------

def _add_readme_tab(wb, season_range: str | None = None) -> None:
    ws = wb.create_sheet("Read-Me")
    ws.sheet_properties.tabColor = "A9A9A9"

    training_desc = (
        f"Two most recent completed seasons ({season_range} training data)"
        if season_range
        else "Two most recent completed seasons"
    )

    content = [
        (1,  "Player Value Engine (PVE) — V1",                   Font(bold=True, size=14)),
        (3,  "COLOR KEY — How to Read This Workbook",             Font(bold=True)),
        (4,  "  Blue font   = Hardcoded input — the only cells you should type in",  Font(color="2F5496")),
        (5,  "  Black font  = Formula using data on this same sheet",                Font(color="000000")),
        (6,  "  Green font  = Formula pulling data from another tab",                Font(color="375623")),
        (7,  "  Red font    = Error, alert, or eligibility warning",                 Font(color="C00000")),
        (9,  "TAB GUIDE (left to right)",                         Font(bold=True)),
        (10, "  Dashboard        — master status flag, budget summary, error checks",          None),
        (11, "  Read-Me          — instructions, color key, and scope notes (this tab)",       None),
        (12, "  Inputs           — the ONLY place to type: budget, win curve, EPA/win, asking prices", None),
        (13, "  Calculations     — step-by-step math for each candidate (auditable; do not edit formulas)", None),
        (14, "  Candidate Board  — full candidate output; blue cells can be overridden directly", None),
        (15, "  Recommended Roster — budget-constrained signing recommendation with live formulas", None),
        (16, "  Depth Comparison — current roster vs. candidates by position with credible intervals", None),
        (17, "  Raw Data         — full projected points added for all FBS players (reference and analysis)", None),
        (19, "COLUMN GUIDE — Candidate Board",                    Font(bold=True)),
        (20, "  Marginal Wins   — projected wins added over the backup this player would displace", None),
        (21, "  Wins 80% Lo/Hi  — 80% credible interval: narrow = high floor, wide = boom-or-bust", None),
        (22, "  Max Bid $       — the most your program should pay before the deal stops penciling out", None),
        (23, "  NIL Ask $       — what the player/agent is asking (pulled from Inputs tab)",   None),
        (24, "  Edge $          — Max Bid minus NIL Ask; positive = good deal, negative = overpay", None),
        (26, "LIVE UPDATES",                                       Font(bold=True)),
        (27, "  Edit anything on the Inputs tab and Max Bid, NIL Ask, Edge, and Recommended Roster all recalculate.", None),
        (28, "  No need to re-run Python for budget, win-curve, EPA/win, or asking-price sensitivity.", None),
        (30, "IMPORTANT — PROJECTIONS ASSUME LAST SEASON'S PLAY VOLUME", Font(bold=True)),
        (31, "  The model projects per-play skill (proj_points_per_play) and multiplies by last season's play count (proj_snaps).", None),
        (32, "  Both columns are on the Raw Data tab, so any projection audits exactly: proj_points_mean = proj_points_per_play × proj_snaps.", None),
        (33, "  If you sign a player into a different role (starter elsewhere → depth here, or backup → starter), rescale by", None),
        (34, "  hand (rate × expected snaps) or enter a usage-adjusted estimate in the Manual Points Added column of your input workbook.", None),
        (36, "SCOPE LIMITS (V1)",                                  Font(bold=True)),
        (37, "  - Offensive skill positions only: QB, RB, WR, TE", None),
        (38, f"  - {training_desc}", None),
        (39, "  - Win-value curve covers regular season only (0-12 wins); CFP not modeled", None),
        (40, "  - Two-deep only; depth beyond the backup is not modeled", None),
        (41, "  - Data Source = insufficient: fewer than ~10 estimated snaps; model cannot price this player.", None),
        (42, "    Provide a manual points-added estimate in your input Candidates tab based on scouting and game film.", None),
        (43, "  - Skill Position Portal Budget covers QB, RB, WR, TE only", None),
        (44, "  - 'user-asserted' source means the value came from your Manual Points Added input, not the model", None),
        (46, "UNCERTAINTY",                                        Font(bold=True)),
        (47, "  Marginal Wins is shown with 80% credible interval bounds.", None),
        (48, "  Wide bands = less data — treat those players with more caution.", None),
        (49, "  Dollar amounts carry the same uncertainty as the projected-wins (WAA) bands.", None),
    ]

    for row, text, font in content:
        cell = ws.cell(row=row, column=1, value=text)
        if font is not None:
            cell.font = font

    ws.column_dimensions["A"].width = 95


# ---------------------------------------------------------------------------
# Tab 3: Inputs
# ---------------------------------------------------------------------------

def _add_inputs_tab(wb, inputs: dict, lambda_risk: float) -> None:
    ws = wb.create_sheet("Inputs")
    ws.sheet_properties.tabColor = "FFC000"

    note_font  = Font(italic=True, color="595959", size=9)
    label_font = Font(bold=True)

    # Row 1: title
    _title_row(ws, 1, "PROGRAM INPUTS — Edit blue cells only", 3, height=22)

    # Row 2: PROGRAM CONFIGURATION
    _section_label(ws, 2, "PROGRAM CONFIGURATION", 3)

    config_rows = [
        (3, "Total Budget ($)",                     inputs["budget"],                 _FMT_DOLLAR,
            "Portal spend ceiling for QB, RB, WR, TE only"),
        (4, "Preseason Win Projection",             int(inputs["preseason_wins"]),     None,
            "Expected wins before any portal signings (0-12)"),
        (5, "Program Name",                         inputs.get("program_name", ""),   None,
            "Used for display only"),
        (6, "Lambda (Risk Aversion — reference)",   lambda_risk,                       None,
            "Set in Python run; 0 = maximize upside, 1 = minimize variance. Re-run pipeline to change."),
    ]
    for r, label, val, fmt, note in config_rows:
        ws.cell(row=r, column=1, value=label).font = label_font
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = _BLUE
        if fmt:
            vc.number_format = fmt
        ws.cell(row=r, column=3, value=note).font = note_font

    # Row 8: EPA PER WIN CONVERSION
    _section_label(ws, 8, "EPA PER WIN CONVERSION", 3)

    ws.cell(row=9, column=1, value="EPA per Win").font = label_font
    epa_c = ws.cell(row=9, column=2, value=EPA_PER_WIN)
    epa_c.font = _BLUE
    ws.cell(row=9, column=3,
        value="How many cumulative EPA points equal one additional win").font = note_font

    epa_notes = [
        (10, "What this controls: Projected WAA (wins) = projected points added / this number. "
             "Lowering it increases every player's projected wins and Max Bid; raising it decreases them."),
        (11, "Example — setting to 100: each player's EPA converts to MORE wins; Max Bid goes UP. "
             "A player with +130 EPA becomes +1.30 wins instead of +1.00."),
        (12, "Example — setting to 200: each player's EPA converts to FEWER wins; Max Bid goes DOWN. "
             "A player with +130 EPA becomes +0.65 wins instead of +1.00."),
    ]
    for r, text in epa_notes:
        c = ws.cell(row=r, column=1, value=text)
        c.font = note_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    # Row 14: WIN VALUE CURVE
    _section_label(ws, 14, "WIN VALUE CURVE", 3)

    # Row 15: column headers
    for c_i, h in enumerate(["Wins Gained (N)", "Dollar Value of Win N ($)", "Notes"], 1):
        cell = ws.cell(row=15, column=c_i, value=h)
        cell.fill = _FILL_DARK_BLUE
        cell.font = Font(color="FFFFFF", bold=True)

    # Rows 16-27: win curve — STABLE POSITIONS (formulas reference these exact rows)
    win_notes = {
        6:  "Bowl eligibility threshold — spike this for bubble programs",
        10: "Conference championship contention",
        12: "Perfect regular season",
    }
    for i, (wins_n, win_val) in enumerate(sorted(inputs["win_curve"].items()), start=16):
        ws.cell(row=i, column=1, value=int(wins_n)).font = _BLUE
        vc = ws.cell(row=i, column=2, value=float(win_val))
        vc.font = _BLUE
        vc.number_format = _FMT_DOLLAR
        wnote = win_notes.get(int(wins_n), "")
        if wnote:
            ws.cell(row=i, column=3, value=wnote).font = note_font

    # Row 29: TWO-DEEP DEPTH CHART
    cell = ws.cell(row=29, column=1, value="TWO-DEEP DEPTH CHART")
    cell.font = Font(bold=True, color="1F3864")
    cell.fill = _FILL_LIGHT_BLUE
    ws.merge_cells(start_row=29, start_column=1, end_row=29, end_column=5)

    # Row 30: depth chart column headers
    for c_i, h in enumerate(["Position", "Starter Name", "Displaced Player", "Starter ID", "Backup ID"], 1):
        cell = ws.cell(row=30, column=c_i, value=h)
        cell.fill = _FILL_DARK_BLUE
        cell.font = Font(color="FFFFFF", bold=True)

    # Rows 31-39: depth chart data
    for i, (_, dr) in enumerate(inputs["depth_chart"].iterrows(), start=31):
        ws.cell(row=i, column=1, value=dr.get("position", "")).font = _BLUE
        ws.cell(row=i, column=2, value=dr.get("starter_name", "")).font = _BLUE
        ws.cell(row=i, column=3, value=dr.get("backup_name", "")).font = _BLUE
        ws.cell(row=i, column=4, value=dr.get("starter_id", "")).font = _BLUE
        ws.cell(row=i, column=5, value=dr.get("backup_id", "")).font = _BLUE

    # Row 40: CANDIDATE ASKING PRICES — STABLE ROW (formulas reference $A$42:$B$100)
    cell = ws.cell(row=40, column=1, value="CANDIDATE ASKING PRICES")
    cell.font = Font(bold=True, color="1F3864")
    cell.fill = _FILL_LIGHT_BLUE
    ws.merge_cells(start_row=40, start_column=1, end_row=40, end_column=2)

    # Row 41: asking prices column headers
    for c_i, h in enumerate(["Player", "NIL Ask $"], 1):
        cell = ws.cell(row=41, column=c_i, value=h)
        cell.fill = _FILL_DARK_BLUE
        cell.font = Font(color="FFFFFF", bold=True)

    # Rows 42+: pre-populated from input candidates (user can edit)
    for i, (_, cand) in enumerate(inputs["candidates"].iterrows(), start=42):
        ws.cell(row=i, column=1, value=cand.get("player", "")).font = _BLUE
        vc = ws.cell(row=i, column=2, value=float(cand.get("asking_price", 0)))
        vc.font = _BLUE
        vc.number_format = _FMT_DOLLAR

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


# ---------------------------------------------------------------------------
# Tab 4: Calculations
# ---------------------------------------------------------------------------

def _add_calculations_tab(wb, candidate_board: pd.DataFrame) -> None:
    ws = wb.create_sheet("Calculations")
    ws.sheet_properties.tabColor = "D9D9D9"

    n_cols = 11

    _title_row(ws, 1, "CALCULATIONS — Intermediate Steps (Auditable)", n_cols)

    calc_hdrs = [
        "Player", "Proj EPA", "Incumbent EPA", "Marginal EPA",
        "EPA/Win", "Marginal Wins", "Preseason Wins", "$/Win",
        "Max Bid $", "NIL Ask $", "Edge $",
    ]
    _hdr(ws, 2, n_cols)
    for c_i, h in enumerate(calc_hdrs, 1):
        ws.cell(row=2, column=c_i).value = h

    for i, (_, row) in enumerate(candidate_board.iterrows(), start=3):
        r = i

        # A: Player (blue — hardcoded)
        ws.cell(row=r, column=1, value=row.get("player", "")).font = _BLUE

        # B: Proj EPA (blue — hardcoded; None when insufficient)
        proj_raw = row.get("proj_points_mean")
        b_val = _safe_na(proj_raw)
        b_val = float(b_val) if b_val is not None else None
        bc = ws.cell(row=r, column=2, value=b_val)
        bc.font = _BLUE
        if b_val is not None:
            bc.number_format = _FMT_EPA

        # C: Incumbent EPA (blue — hardcoded)
        inc_raw = row.get("incumbent_points")
        c_val = _safe_na(inc_raw)
        c_val = float(c_val) if c_val is not None else None
        cc = ws.cell(row=r, column=3, value=c_val)
        cc.font = _BLUE
        if c_val is not None:
            cc.number_format = _FMT_EPA

        # D: Marginal EPA = IFERROR(B-C,"") — black (same-sheet formula)
        dc = ws.cell(row=r, column=4, value=f'=IFERROR(B{r}-C{r},"")')
        dc.font = _BLACK
        dc.number_format = _FMT_EPA

        # E: EPA per Win = Inputs!$B$9 — green (cross-sheet)
        ec = ws.cell(row=r, column=5, value=f"={_EPA_WIN_REF}")
        ec.font = _GREEN

        # F: Marginal Wins = IFERROR(D/E,"") — green (cross-sheet via E)
        fc = ws.cell(row=r, column=6, value=f'=IFERROR(D{r}/E{r},"")')
        fc.font = _GREEN
        fc.number_format = _FMT_WINS

        # G: Preseason Wins = Inputs!$B$4 — green (cross-sheet)
        gc = ws.cell(row=r, column=7, value=f"={_PRESEASON_REF}")
        gc.font = _GREEN

        # H: $/Win = VLOOKUP(MIN(G+1,12), win curve, 2) — green (cross-sheet)
        hc = ws.cell(row=r, column=8,
            value=f"=IFERROR(VLOOKUP(MIN(G{r}+1,12),{_WIN_CURVE_REF},2,FALSE),0)")
        hc.font = _GREEN
        hc.number_format = _FMT_DOLLAR

        # I: Max Bid $ = IF(ISNUMBER(F),F*H,"") — black (same-sheet)
        ic = ws.cell(row=r, column=9, value=f'=IF(ISNUMBER(F{r}),F{r}*H{r},"")')
        ic.font = _BLACK
        ic.number_format = _FMT_DOLLAR

        # J: NIL Ask $ = VLOOKUP(player name, asking prices table, 2) — green (cross-sheet)
        jc = ws.cell(row=r, column=10,
            value=f"=IFERROR(VLOOKUP($A{r},{_ASK_REF},2,FALSE),0)")
        jc.font = _GREEN
        jc.number_format = _FMT_DOLLAR

        # K: Edge $ = IF(ISNUMBER(I),I-J,"") — black (same-sheet)
        kc = ws.cell(row=r, column=11, value=f'=IF(ISNUMBER(I{r}),I{r}-J{r},"")')
        kc.font = _BLACK
        kc.number_format = _FMT_DOLLAR

    col_widths = {
        "A": 24, "B": 12, "C": 14, "D": 13, "E": 10,
        "F": 14, "G": 16, "H": 12, "I": 14, "J": 14, "K": 12,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Tab 5: Candidate Board
# ---------------------------------------------------------------------------

def _add_candidate_board_tab(wb, candidate_board: pd.DataFrame,
                             projection_season: int) -> None:
    ws = wb.create_sheet("Candidate Board")
    ws.sheet_properties.tabColor = "70AD47"

    cb_headers = [
        "Player", "Position", "Current Team", "Class", "Data Source",
        "Marginal Wins", "Wins 80% Lo", "Wins 80% Hi",
        "Max Bid $", "NIL Ask $", "Edge $",
        "Eligibility Note", "Proj WAA (wins)",
    ]
    # Column map (1-based):
    #   F=6  Marginal Wins   G=7  Wins 80% Lo   H=8  Wins 80% Hi
    #   I=9  Max Bid $       J=10 NIL Ask $      K=11 Edge $
    #   L=12 Eligibility     M=13 Proj WAA (wins) — Depth Comparison VLOOKUPs col 13

    ws.append(cb_headers)
    _hdr(ws, 1, len(cb_headers))

    CALC_REF = "'Calculations'!$A:$K"

    for row_idx, (_, row) in enumerate(candidate_board.iterrows(), start=2):
        ci_lo_raw = row.get("waa_ci_80_lo")
        ci_lo = float(ci_lo_raw) if _safe_na(ci_lo_raw) is not None else None
        ci_hi_raw = row.get("waa_ci_80_hi")
        ci_hi = float(ci_hi_raw) if _safe_na(ci_hi_raw) is not None else None
        proj_raw = row.get("proj_points_mean")
        proj_wins = (float(proj_raw) / EPA_PER_WIN
                     if _safe_na(proj_raw) is not None else None)

        year_class = row.get("year_class", "UNK") or "UNK"
        elig_note = (
            f"Verify {projection_season} eligibility — was {year_class} in training data"
            if year_class in ELIGIBILITY_WARN_CLASSES
            else ""
        )

        # A-E: hardcoded (blue)
        ws.cell(row=row_idx, column=1, value=row.get("player")).font = _BLUE
        ws.cell(row=row_idx, column=2, value=row.get("position")).font = _BLUE
        ws.cell(row=row_idx, column=3, value=row.get("current_team")).font = _BLUE
        ws.cell(row=row_idx, column=4, value=year_class).font = _BLUE
        ws.cell(row=row_idx, column=5, value=row.get("source", "model")).font = _BLUE

        # F: Marginal Wins — VLOOKUP from Calculations col 6 (green)
        fc = ws.cell(row=row_idx, column=6,
            value=f'=IFERROR(VLOOKUP($A{row_idx},{CALC_REF},6,FALSE),"")')
        fc.font = _GREEN
        fc.number_format = _FMT_WINS

        # G: Wins 80% Lo (blue — hardcoded CI from Python)
        gc = ws.cell(row=row_idx, column=7, value=ci_lo)
        gc.font = _BLUE
        if ci_lo is not None:
            gc.number_format = _FMT_WINS

        # H: Wins 80% Hi (blue — hardcoded CI from Python)
        hc = ws.cell(row=row_idx, column=8, value=ci_hi)
        hc.font = _BLUE
        if ci_hi is not None:
            hc.number_format = _FMT_WINS

        # I: Max Bid $ — VLOOKUP from Calculations col 9 (green)
        ic = ws.cell(row=row_idx, column=9,
            value=f'=IFERROR(VLOOKUP($A{row_idx},{CALC_REF},9,FALSE),"")')
        ic.font = _GREEN
        ic.number_format = _FMT_DOLLAR

        # J: NIL Ask $ — VLOOKUP from Calculations col 10 (green)
        jc = ws.cell(row=row_idx, column=10,
            value=f'=IFERROR(VLOOKUP($A{row_idx},{CALC_REF},10,FALSE),"")')
        jc.font = _GREEN
        jc.number_format = _FMT_DOLLAR

        # K: Edge $ — VLOOKUP from Calculations col 11 (green)
        kc = ws.cell(row=row_idx, column=11,
            value=f'=IFERROR(VLOOKUP($A{row_idx},{CALC_REF},11,FALSE),"")')
        kc.font = _GREEN
        kc.number_format = _FMT_DOLLAR

        # L: Eligibility Note (red when warning, else blue)
        lc = ws.cell(row=row_idx, column=12, value=elig_note)
        lc.font = _RED if elig_note else _BLUE

        # M: Proj WAA (wins) — absolute point estimate (blue)
        mc = ws.cell(row=row_idx, column=13, value=proj_wins)
        mc.font = _BLUE
        if proj_wins is not None:
            mc.number_format = _FMT_WINS

    for col in range(1, len(cb_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


# ---------------------------------------------------------------------------
# Tab 6: Recommended Roster
# ---------------------------------------------------------------------------

def _add_recommended_roster_tab(wb, candidate_board: pd.DataFrame) -> None:
    ws = wb.create_sheet("Recommended Roster")
    ws.sheet_properties.tabColor = "70AD47"

    # Sort by Python surplus descending to establish greedy priority order.
    cb_sorted = candidate_board.copy()
    cb_sorted["_sort_key"] = pd.to_numeric(
        cb_sorted.get("surplus", None), errors="coerce"
    ).fillna(-1e9)
    cb_sorted = cb_sorted.sort_values("_sort_key", ascending=False).drop("_sort_key", axis=1)

    n_cands   = len(cb_sorted)
    DATA_START = 4
    DATA_END   = DATA_START + n_cands - 1
    CB_REF     = "'Candidate Board'!$A:$M"

    note_font = Font(italic=True, color="595959")
    bold_font = Font(bold=True)

    # Row 1: title
    ws.cell(row=1, column=1,
        value="RECOMMENDED ROSTER — Updates automatically when you edit the Inputs tab"
    ).font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    # Row 2: legend
    ws.cell(row=2, column=1,
        value="Green = recommended signing  |  Yellow = positive value but over budget  |  No highlight = not worth signing"
    ).font = note_font
    ws.merge_cells("A2:G2")

    # Row 3: column headers
    rr_hdrs = ["Player", "Position", "Marginal Wins", "Asking $", "Edge $",
               "Cum. Spend Before", "Recommended?"]
    for c_i, h in enumerate(rr_hdrs, 1):
        c = ws.cell(row=3, column=c_i, value=h)
        c.fill = _FILL_DARK_BLUE
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["F"].hidden = True

    for i, (_, row) in enumerate(cb_sorted.iterrows()):
        r = DATA_START + i
        asking = float(row.get("asking_price", 0))

        # F: cumulative spend before this player (greedy chain)
        cum_f = 0 if i == 0 else f'=F{r-1}+IF(G{r-1}="Yes",D{r-1},0)'

        # G: recommend if edge > 0 and fits remaining budget
        sel_f = f'=IF(AND(ISNUMBER(E{r}),E{r}>0,D{r}+F{r}<={_BUDGET_REF}),"Yes","No")'

        ws.cell(row=r, column=1, value=row.get("player")).font = _BLUE
        ws.cell(row=r, column=2, value=row.get("position")).font = _BLUE

        # C: Marginal Wins — VLOOKUP from Candidate Board col 6 (green)
        cc = ws.cell(row=r, column=3,
            value=f'=IFERROR(VLOOKUP(A{r},{CB_REF},6,FALSE),"")')
        cc.font = _GREEN
        cc.number_format = _FMT_WINS

        # D: Asking $ (blue — hardcoded)
        dc = ws.cell(row=r, column=4, value=asking)
        dc.font = _BLUE
        dc.number_format = _FMT_DOLLAR

        # E: Edge $ — VLOOKUP from Candidate Board col 11 (green)
        ec = ws.cell(row=r, column=5,
            value=f'=IFERROR(VLOOKUP(A{r},{CB_REF},11,FALSE),"")')
        ec.font = _GREEN
        ec.number_format = _FMT_DOLLAR

        # F: hidden running-total helper (black — same-sheet)
        fc = ws.cell(row=r, column=6, value=cum_f)
        fc.font = _BLACK
        fc.number_format = _FMT_DOLLAR

        # G: Recommended? (green — references Inputs via budget ref)
        gc = ws.cell(row=r, column=7, value=sel_f)
        gc.font = _GREEN

    # Conditional formatting
    data_range  = f"A{DATA_START}:G{DATA_END}"
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'$G{DATA_START}="Yes"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'AND($G{DATA_START}="No",ISNUMBER($E{DATA_START}),$E{DATA_START}>0)'],
            fill=yellow_fill,
        ),
    )

    # Summary block
    summ = DATA_END + 2
    ws.cell(row=summ,   column=1, value="Total Spend").font = bold_font
    ts = ws.cell(row=summ, column=2,
        value=f'=SUMPRODUCT((G{DATA_START}:G{DATA_END}="Yes")*D{DATA_START}:D{DATA_END})')
    ts.number_format = _FMT_DOLLAR
    ts.font = _BLACK

    ws.cell(row=summ+1, column=1, value="Budget").font = bold_font
    bc = ws.cell(row=summ+1, column=2, value=f"={_BUDGET_REF}")
    bc.number_format = _FMT_DOLLAR
    bc.font = _GREEN

    ws.cell(row=summ+2, column=1, value="Budget Remaining").font = bold_font
    rc = ws.cell(row=summ+2, column=2, value=f"=B{summ+1}-B{summ}")
    rc.number_format = _FMT_DOLLAR
    rc.font = _BLACK

    ws.cell(row=summ+4, column=1, value="Positional Allocation").font = bold_font
    for j, pos in enumerate(["QB", "RB", "WR", "TE"]):
        pos_row = summ + 5 + j
        ws.cell(row=pos_row, column=1, value=pos)
        pc = ws.cell(row=pos_row, column=2,
            value=(f'=SUMPRODUCT((G{DATA_START}:G{DATA_END}="Yes")'
                   f'*(B{DATA_START}:B{DATA_END}="{pos}")'
                   f'*D{DATA_START}:D{DATA_END})'))
        pc.number_format = _FMT_DOLLAR
        pc.font = _GREEN

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["G"].width = 14


# ---------------------------------------------------------------------------
# Tab 7: Depth Comparison
# ---------------------------------------------------------------------------

def _add_depth_comparison_tab(
    wb,
    candidate_board: pd.DataFrame,
    depth_chart: pd.DataFrame,
) -> None:
    """Add a Depth Comparison tab showing current roster vs. portal candidate WAA.

    Candidate rows use VLOOKUP formulas keyed on player name so edits to the
    Candidate Board (cols F/G/H/M) propagate here without re-running Python.
    Starter/backup rows are static numbers editable directly in this tab.

    Candidate Board column map referenced by VLOOKUP:
      F=6  Marginal Wins (vs backup)
      G=7  Wins 80% Lo (absolute)
      H=8  Wins 80% Hi (absolute)
      M=13 Proj WAA (wins) — absolute point estimate

    depth_chart must have: position, starter_name, backup_name,
    starter_wins, backup_wins (None when no player_id; treated as 0-WAA).
    """
    ws = wb.create_sheet("Depth Comparison")
    ws.sheet_properties.tabColor = "70AD47"

    # Candidate Board lookup range — spans all 13 columns A:M
    CB_REF      = "'Candidate Board'!$A:$M"
    CB_MARG_COL = 6    # F: Marginal Wins
    CB_LO_COL   = 7    # G: Wins 80% Lo (absolute)
    CB_HI_COL   = 8    # H: Wins 80% Hi (absolute)
    CB_PROJ_COL = 13   # M: Proj WAA (wins) absolute

    def _vl(name_cell: str, col_idx: int) -> str:
        return f'=IFERROR(VLOOKUP({name_cell},{CB_REF},{col_idx},FALSE),"—")'

    pos_order  = ["QB", "RB", "WR", "TE"]
    pos_labels = {"QB": "QUARTERBACK", "RB": "RUNNING BACK", "WR": "WIDE RECEIVER", "TE": "TIGHT END"}
    pos_hdr_color = {"QB": "2E4057", "RB": "1E5C27", "WR": "7B2D00", "TE": "3D1A5C"}

    title_font   = Font(bold=True, size=13)
    note_font    = Font(italic=True, color="595959")
    bold_font    = Font(bold=True)
    sub_hdr_font = Font(bold=True, size=9, color="595959")
    purple_font  = Font(italic=True, bold=True, color="7030A0")
    live_note    = Font(italic=True, size=9, color="2E75B6")

    starter_fill = PatternFill("solid", fgColor="DDEEFF")
    backup_fill  = PatternFill("solid", fgColor="F2F2F2")
    green_fill   = PatternFill("solid", fgColor="E2EFDA")
    pos_fills    = {k: PatternFill("solid", fgColor=v) for k, v in pos_hdr_color.items()}

    FMT_WINS  = '+0.00;-0.00;"-"'
    FMT_RANGE = '0.00'
    N_COLS = 7  # A-G

    # Row 1: title
    ws.cell(row=1, column=1,
            value="Depth Chart Comparison — Projected Wins & Credible Intervals"
            ).font = title_font
    ws.merge_cells("A1:G1")

    # Row 2: CI explanation
    ws.cell(row=2, column=1,
            value=(
                "Proj Wins (WAA) = absolute projected points added / 130.  Positive = above-average at that position. "
                "80% Low / 80% High = range that covers the true outcome 80% of the time. "
                "Narrow range = high floor; wide range = boom-or-bust. "
                "vs Backup = marginal uplift over the weakest displaced player."
            )).font = note_font
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 30

    # Row 3: live-update note
    ws.cell(row=3, column=1,
            value=(
                "LIVE: Edit Proj Wins, 80% Low/High, or vs Backup directly in the Candidate Board — "
                "this tab updates automatically. "
                "Starter/backup wins (blue/gray rows) can be edited directly here."
            )).font = live_note
    ws.merge_cells("A3:G3")

    all_data_start = None
    all_data_end   = None

    row = 5  # first data row

    for pos in pos_order:
        dc_pos    = depth_chart[depth_chart["position"] == pos]
        cands_pos = candidate_board[candidate_board["position"] == pos].copy()

        if dc_pos.empty and cands_pos.empty:
            continue

        bw_vals    = [v for v in dc_pos.get("backup_wins", pd.Series()).tolist()
                      if v is not None and pd.notna(v)]
        weakest_bw = min(bw_vals) if bw_vals else 0.0

        cands_pos["_sort"] = pd.to_numeric(
            cands_pos.get("marginal_mean_wins"), errors="coerce"
        ).fillna(-1e9)
        cands_pos = cands_pos.sort_values("_sort", ascending=False).drop("_sort", axis=1)

        if all_data_start is None:
            all_data_start = row

        # Position header
        fill = pos_fills.get(pos, PatternFill("solid", fgColor="2E4057"))
        ws.cell(row=row, column=1, value=pos_labels.get(pos, pos))
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        for c in range(2, N_COLS + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.merge_cells(f"A{row}:G{row}")
        row += 1

        # Column sub-header
        for c_i, h in enumerate(
            ["Role", "Player", "Proj Wins", "80% Low", "80% High", "Range", "vs Backup"], 1
        ):
            ws.cell(row=row, column=c_i, value=h).font = sub_hdr_font
        row += 1

        # Depth chart slots
        n_slots = len(dc_pos)
        for slot_i, (_, dr) in enumerate(dc_pos.iterrows(), start=1):
            slot_sfx = f" (Slot {slot_i})" if n_slots > 1 else ""

            ws.cell(row=row, column=1, value=f"Starter{slot_sfx}")
            ws.cell(row=row, column=2, value=dr.get("starter_name") or "—")
            sw = dr.get("starter_wins")
            if sw is not None and pd.notna(sw):
                ws.cell(row=row, column=3, value=float(sw)).number_format = FMT_WINS
            else:
                ws.cell(row=row, column=3, value="—")
            for c in [4, 5, 6, 7]:
                ws.cell(row=row, column=c, value="—")
            for c in range(1, N_COLS + 1):
                ws.cell(row=row, column=c).fill = starter_fill
            row += 1

            ws.cell(row=row, column=1, value=f"Backup{slot_sfx}")
            ws.cell(row=row, column=2, value=dr.get("backup_name") or "—")
            bw = dr.get("backup_wins")
            if bw is not None and pd.notna(bw):
                ws.cell(row=row, column=3, value=float(bw)).number_format = FMT_WINS
            else:
                ws.cell(row=row, column=3, value="—")
            for c in [4, 5, 6]:
                ws.cell(row=row, column=c, value="—")
            bw_null    = bw is None or (isinstance(bw, float) and pd.isna(bw))
            is_weakest = not bw_null and abs(float(bw) - weakest_bw) < 0.001
            is_zero_tgt = bw_null and weakest_bw <= 0.0
            if is_weakest or is_zero_tgt:
                ws.cell(row=row, column=7, value="← displacement target").font = purple_font
            else:
                ws.cell(row=row, column=7, value="—")
            for c in range(1, N_COLS + 1):
                ws.cell(row=row, column=c).fill = backup_fill
            row += 1

        # Portal Candidates separator
        ws.cell(row=row, column=1, value="Portal Candidates ↓").font = Font(bold=True, size=9, color="595959")
        row += 1

        # Candidate rows — VLOOKUP formulas keyed on player name
        for _, cand in cands_pos.iterrows():
            pname = cand.get("player")
            bcell = f"$B{row}"

            ws.cell(row=row, column=1, value="Candidate")
            ws.cell(row=row, column=2, value=pname)
            ws.cell(row=row, column=3, value=_vl(bcell, CB_PROJ_COL)).number_format = FMT_WINS
            ws.cell(row=row, column=4, value=_vl(bcell, CB_LO_COL)).number_format = FMT_WINS
            ws.cell(row=row, column=5, value=_vl(bcell, CB_HI_COL)).number_format = FMT_WINS
            ws.cell(row=row, column=6,
                    value=(f'=IFERROR('
                           f'VLOOKUP({bcell},{CB_REF},{CB_HI_COL},FALSE)'
                           f'-VLOOKUP({bcell},{CB_REF},{CB_LO_COL},FALSE)'
                           f',"—")')
                    ).number_format = FMT_RANGE
            ws.cell(row=row, column=7, value=_vl(bcell, CB_MARG_COL)).number_format = FMT_WINS

            row += 1

        all_data_end = row - 1
        row += 1  # spacer

    # Conditional formatting: green row when vs Backup > 0
    if all_data_start is not None and all_data_end is not None:
        ws.conditional_formatting.add(
            f"A{all_data_start}:G{all_data_end}",
            FormulaRule(
                formula=[f'=AND($A{all_data_start}="Candidate",'
                         f'ISNUMBER($G{all_data_start}),$G{all_data_start}>0)'],
                fill=green_fill,
            ),
        )

    # Portfolio summary
    row += 1
    ws.cell(row=row, column=1, value="Portfolio Summary").font = Font(bold=True, size=11)
    row += 1

    pos_surp  = pd.to_numeric(candidate_board.get("surplus", pd.Series(dtype=float)), errors="coerce")
    marg_all  = pd.to_numeric(candidate_board.get("marginal_mean_wins", pd.Series(dtype=float)), errors="coerce")
    pos_mask  = (pos_surp > 0) & (marg_all > 0)
    total_marg = float(marg_all[pos_mask].dropna().sum())

    ci_lo_all = pd.to_numeric(candidate_board.get("waa_ci_80_lo", pd.Series(dtype=float)), errors="coerce")
    ci_hi_all = pd.to_numeric(candidate_board.get("waa_ci_80_hi", pd.Series(dtype=float)), errors="coerce")
    total_lo = total_hi = None
    if "incumbent_points" in candidate_board.columns:
        inc_wins = pd.to_numeric(candidate_board["incumbent_points"], errors="coerce") / EPA_PER_WIN
        lo_vals  = (ci_lo_all - inc_wins)[pos_mask].dropna()
        hi_vals  = (ci_hi_all - inc_wins)[pos_mask].dropna()
        if not lo_vals.empty:
            total_lo = float(lo_vals.sum())
            total_hi = float(hi_vals.sum())

    ws.cell(row=row, column=1,
            value="Expected marginal wins (positive-edge candidates):").font = bold_font
    c_tot = ws.cell(row=row, column=2, value=round(total_marg, 2))
    c_tot.number_format = '+0.00;-0.00;"-"'
    c_tot.font = Font(bold=True, color="2E75B6")

    if total_lo is not None and total_hi is not None:
        ws.cell(row=row, column=3, value="80% range:").font = note_font
        lo_c = ws.cell(row=row, column=4, value=round(total_lo, 2))
        lo_c.number_format = '+0.00;-0.00;"-"'
        lo_c.font = note_font
        ws.cell(row=row, column=5, value="to").font = note_font
        hi_c = ws.cell(row=row, column=6, value=round(total_hi, 2))
        hi_c.number_format = '+0.00;-0.00;"-"'
        hi_c.font = note_font
    row += 1

    ws.cell(row=row, column=1,
            value="(Maximum possible uplift if all positive-edge candidates signed.)"
            ).font = note_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 14


# ---------------------------------------------------------------------------
# Tab 8: Raw Data
# ---------------------------------------------------------------------------

def _add_raw_data_tab(wb, projections: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw Data")
    ws.sheet_properties.tabColor = "595959"

    raw_cols = [
        "player_id", "player", "team", "conference", "position",
        "data_status", "year_class", "is_transfer",
        "proj_snaps", "proj_points_per_play",
        "proj_points_mean", "proj_points_sd", "proj_points_ci_80_lo", "proj_points_ci_80_hi",
        "proj_points_ci_95_lo", "proj_points_ci_95_hi",
    ]
    available_cols = [c for c in raw_cols if c in projections.columns]
    n_cols = len(available_cols)

    _title_row(ws, 1, "RAW DATA — Full Projected Points Added (All FBS Players, Reference Only)", n_cols)

    note = ws.cell(row=2, column=1,
        value="Reference and advanced analysis only. Do not edit. Source: projected_points.parquet")
    note.font = Font(italic=True, color="595959")
    if n_cols > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    _hdr(ws, 3, n_cols)
    for c_i, h in enumerate(available_cols, 1):
        ws.cell(row=3, column=c_i).value = h

    ws.freeze_panes = "A4"

    epa_cols  = {"proj_points_mean", "proj_points_ci_80_lo", "proj_points_ci_80_hi", "proj_points_ci_95_lo", "proj_points_ci_95_hi"}
    sd_cols   = {"proj_points_sd"}
    rate_cols = {"proj_points_per_play"}
    int_cols  = {"proj_snaps"}

    sub_df = projections[available_cols].copy()
    for row_idx, (_, row) in enumerate(sub_df.iterrows(), start=4):
        for c_i, col_name in enumerate(available_cols, 1):
            val = row[col_name]
            val = _safe_na(val)
            cell = ws.cell(row=row_idx, column=c_i, value=val)
            cell.font = _BLUE
            if col_name in epa_cols and val is not None:
                cell.number_format = _FMT_EPA
            elif col_name in sd_cols and val is not None:
                cell.number_format = "0.0"
            elif col_name in rate_cols and val is not None:
                cell.number_format = "+0.000;-0.000"
            elif col_name in int_cols and val is not None:
                cell.number_format = "0"

    col_widths = {
        "player_id": 14, "player": 22, "team": 20, "conference": 18,
        "position": 10, "data_status": 14, "year_class": 12, "is_transfer": 12,
        "proj_snaps": 12, "proj_points_per_play": 14,
        "proj_points_mean": 12, "proj_points_sd": 10,
        "proj_points_ci_80_lo": 12, "proj_points_ci_80_hi": 12,
        "proj_points_ci_95_lo": 12, "proj_points_ci_95_hi": 12,
    }
    for c_i, col_name in enumerate(available_cols, 1):
        ws.column_dimensions[get_column_letter(c_i)].width = col_widths.get(col_name, 12)


# ---------------------------------------------------------------------------
# Main output writer
# ---------------------------------------------------------------------------

def write_output_workbook(
    output_path: str | Path,
    candidate_board: pd.DataFrame,
    selected_df: pd.DataFrame,
    total_spend: float,
    inputs: dict,
    projections: pd.DataFrame,
    lambda_risk: float = 0.5,
    projection_season: int | None = None,
    season_range: str | None = None,
) -> None:
    """Write the 8-tab FAST-standard report workbook.

    Tab order (left to right):
      Dashboard, Read-Me, Inputs, Calculations, Candidate Board,
      Recommended Roster, Depth Comparison, Raw Data

    projection_season / season_range: the model's projection target year and
    training window (from the projection metadata sidecar). Used for the
    eligibility note and the Read-Me scope line so no year is hardcoded. Fall
    back to the current calendar year / a generic phrase when not provided.

    Stable Inputs tab cell addresses (formulas in Calculations depend on these):
      B3  = Total Budget
      B4  = Preseason Win Projection
      B9  = EPA per Win
      A16:B27 = Win Value Curve
      A42:B100 = Candidate Asking Prices lookup table
    """
    if projection_season is None:
        from datetime import datetime
        projection_season = datetime.now().year

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_dashboard_tab(wb)
    _add_readme_tab(wb, season_range)
    _add_inputs_tab(wb, inputs, lambda_risk)
    _add_calculations_tab(wb, candidate_board)
    _add_candidate_board_tab(wb, candidate_board, projection_season)
    _add_recommended_roster_tab(wb, candidate_board)
    _add_depth_comparison_tab(wb, candidate_board, inputs["depth_chart"])
    _add_raw_data_tab(wb, projections)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Report written to {output_path}")

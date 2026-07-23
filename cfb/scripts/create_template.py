"""Generate templates/input_template.xlsx — blank team input workbook.

Run once to (re)create the template:
    python scripts/create_template.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(color="FFFFFF", bold=True)
LABEL_FONT    = Font(bold=True)
NOTE_FONT     = Font(italic=True, color="595959")
THIN_BORDER   = Border(
    bottom=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
)


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _note(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT


def _set_col_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Tab 1: Config ────────────────────────────────────────────────────────────
def _build_config(ws) -> None:
    ws.title = "Config"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45

    ws["A1"] = "Player Value Engine — Team Config"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:C1")

    rows = [
        ("Setting",               "Value",     "Notes"),
        ("Skill Position Portal Budget", "",    "Portal spend for QB, RB, WR, and TE ONLY. Do not include OL, defense, or special teams — this tool models offense only."),
        ("Preseason Win Projection", "",       "Expected wins with current roster before any signings (integer 0–12)"),
        ("Program Name",          "",          "Used for labeling in the output workbook"),
        ("Risk Aversion (lambda)","0.5",       "Optimizer risk param: 0=risk-neutral, 1=very risk-averse. Default 0.5."),
    ]
    _header_row(ws, 2, rows[0])
    for r, (a, b, c) in enumerate(rows[1:], 3):
        ws.cell(row=r, column=1, value=a).font = LABEL_FONT
        ws.cell(row=r, column=2, value=b)
        _note(ws, r, 3, c)

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"


# ── Tab 2: Win Value Curve ────────────────────────────────────────────────────
def _build_win_curve(ws) -> None:
    ws.title = "Win Value Curve"
    _set_col_widths(ws, {1: 14, 2: 20, 3: 55})

    ws["A1"] = "Win Value Curve"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:C1")

    ws["A2"] = "Instructions"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = "Enter the dollar value YOUR program places on going from N-1 to N wins."
    ws["B2"].font = NOTE_FONT
    ws.merge_cells("B2:C2")

    ws["A3"] = "Scope"
    ws["A3"].font = LABEL_FONT
    ws["B3"] = "Regular-season wins only (0–12). CFP/conference championship not modeled in V1."
    ws["B3"].font = NOTE_FONT
    ws.merge_cells("B3:C3")

    ws["A4"] = "Editability"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = "Spike any win that has outsized importance (e.g., win 6 = bowl eligibility = job security)."
    ws["B4"].font = NOTE_FONT
    ws.merge_cells("B4:C4")

    _header_row(ws, 6, ["Wins Gained (N)", "Dollar Value of Win N ($)", "Notes"])

    examples = {
        1:  "Early non-conf wins — low financial impact for most programs",
        2:  "",
        3:  "",
        4:  "",
        5:  "",
        6:  "Bowl eligibility threshold — spike this for bubble programs",
        7:  "",
        8:  "Conference championship game contention — most conferences select top records, no divisions",
        9:  "",
        10: "Conference championship contention",
        11: "Likely conference champion",
        12: "Perfect regular season — rare, extremely high value for contenders",
    }
    for row_i, (wins, note) in enumerate(examples.items(), 7):
        ws.cell(row=row_i, column=1, value=wins)
        ws.cell(row=row_i, column=2, value="")   # user fills in
        _note(ws, row_i, 3, note)

    ws.freeze_panes = "A7"


# ── Tab 3: Two-Deep ───────────────────────────────────────────────────────────
def _build_two_deep(ws) -> None:
    ws.title = "Two-Deep"
    _set_col_widths(ws, {1: 12, 2: 22, 3: 22, 4: 18, 5: 18, 6: 45})

    ws["A1"] = "Two-Deep Depth Chart"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Instructions"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = ("Each row is one position SLOT. 'Starter' = current occupant. "
                "'Displaced Player' = who a transfer signing would push out of that slot. "
                "Add multiple rows for the same position to model multiple slots (e.g., two WR slots). "
                "Use player_id from projected_points.parquet — leave blank to treat as 0-value (replacement-level).")
    ws["B2"].font = NOTE_FONT
    ws.merge_cells("B2:F2")

    _header_row(ws, 4, ["Position", "Starter Name", "Displaced Player",
                        "Starter player_id", "Displaced player_id", "Notes"])

    positions = [
        ("QB",  "", "", "", "", "Slot 1: starter QB — displaced player is current backup"),
        ("QB",  "", "", "", "", "Slot 2: backup QB — displaced player is current 3rd string"),
        ("RB",  "", "", "", "", "Slot 1: lead back — displaced player is current backup"),
        ("RB",  "", "", "", "", "Slot 2: backup RB — displaced player is current 3rd string"),
        ("WR",  "", "", "", "", "Slot 1: WR1 — displaced player is current WR1 backup"),
        ("WR",  "", "", "", "", "Slot 2: WR2 — displaced player is current WR2 backup"),
        ("WR",  "", "", "", "", "Slot 3: WR3 — displaced player is current WR3 backup"),
        ("TE",  "", "", "", "", "Slot 1: starting TE — displaced player is current backup"),
        ("TE",  "", "", "", "", "Slot 2: backup TE — displaced player is current 3rd string"),
    ]
    for r, row in enumerate(positions, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6:
                cell.font = NOTE_FONT

    ws.freeze_panes = "A5"


# ── Tab 4: Candidates ─────────────────────────────────────────────────────────
def _build_candidates(ws) -> None:
    ws.title = "Candidates"
    _set_col_widths(ws, {1: 18, 2: 22, 3: 12, 4: 20, 5: 18, 6: 20, 7: 45})

    ws["A1"] = "Portal Candidate Board"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")

    ws["A2"] = "Instructions"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = ("One row per candidate. player_id from projected_points.parquet. "
                "asking_price in dollars. "
                "manual_points: your own points-added (EPA) estimate — cumulative "
                "season points, same scale as the model (user-asserted, no model CI). "
                "Required for players flagged INSUFFICIENT DATA; also use it when "
                "a player's role will differ from last season (the model assumes "
                "they repeat last season's play volume).")
    ws["B2"].font = NOTE_FONT
    ws.merge_cells("B2:G2")

    _header_row(ws, 4, ["player_id", "Player Name", "Position",
                        "Current Team", "Asking Price ($)", "Manual Points Added (EPA)",
                        "Notes"])

    # Five blank rows for user to fill in
    for r in range(5, 15):
        for c in range(1, 7):
            ws.cell(row=r, column=c, value="")
        _note(ws, r, 7, "")

    ws.freeze_panes = "A5"


# ── Tab 5: Read-Me ─────────────────────────────────────────────────────────────
def _build_readme(ws) -> None:
    ws.title = "Read-Me"
    ws.column_dimensions["A"].width = 90

    ws["A1"] = "Player Value Engine (PVE) — Input Workbook Guide"
    ws["A1"].font = Font(bold=True, size=14)

    lines = [
        ("", ""),
        ("WHAT THIS WORKBOOK DOES", True),
        ("Fill in the four tabs (Config, Win Value Curve, Two-Deep, Candidates),", False),
        ("then run:  python scripts/run_decision.py --input <this_file> --output report.xlsx", False),
        ("", ""),
        ("TAB: Config", True),
        ("  Budget Ceiling       — your total portal spend ceiling in dollars", False),
        ("  Preseason Win Projection — integer 0–12, your expected wins before signings", False),
        ("  Risk Aversion (lambda) — 0.0 = maximize upside; 1.0 = minimize variance", False),
        ("", ""),
        ("TAB: Win Value Curve", True),
        ("  Enter what going from N-1 to N wins is worth to YOUR program.", False),
        ("  There is no right answer — this is YOUR reservation value, not the market's.", False),
        ("  Spike win 6 if bowl eligibility matters. Spike win 10–12 if you're a contender.", False),
        ("  The curve covers regular-season wins 0–12 only. CFP not modeled in V1.", False),
        ("", ""),
        ("TAB: Two-Deep", True),
        ("  Enter starter and backup for each position slot.", False),
        ("  A candidate displaces the backup — marginal value = candidate points minus backup points.", False),
        ("  player_id comes from projected_points.parquet (run train-engine to generate it).", False),
        ("  Leave player_id blank to treat the incumbent as a 0-value (replacement-level) player.", False),
        ("", ""),
        ("TAB: Candidates", True),
        ("  One row per portal candidate you're evaluating.", False),
        ("  player_id must match projected_points.parquet for the model to price them.", False),
        ("  INSUFFICIENT DATA: if a player has no model estimate, enter your own points-added", False),
        ("  (EPA) estimate in the Manual Points Added column.", False),
        ("  ROLE CHANGES: projections assume the player repeats last season's play volume —", False),
        ("  if you're signing them into a bigger or smaller role, enter a usage-adjusted Manual Points Added.", False),
        ("  Manual Points Added rows are labeled 'user-asserted' in the output and carry no CI.", False),
        ("", ""),
        ("IMPORTANT LIMITATIONS (V1)", True),
        ("  - Offensive skill positions only: QB, RB, WR, TE. No defense.", False),
        ("  - Win-value curve excludes CFP and conference championship games.", False),
        ("  - Player value is cumulative points added (EPA); WAA (Wins Above Average) = points / 130.", False),
        ("  - Snap counts are estimated from involvement rate; players with very few plays get", False),
        ("    wide uncertainty bands automatically. Manual Points Added is only needed for INSUFFICIENT", False),
        ("    players and role changes — not for every low-snap player.", False),
        ("  - This is your reservation value, not the market price.", False),
    ]

    for r, (text, bold) in enumerate(lines, 2):
        cell = ws.cell(row=r, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)


def main() -> None:
    TEMPLATES_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_config(wb.create_sheet("Config"))
    _build_win_curve(wb.create_sheet("Win Value Curve"))
    _build_two_deep(wb.create_sheet("Two-Deep"))
    _build_candidates(wb.create_sheet("Candidates"))
    _build_readme(wb.create_sheet("Read-Me"))

    out = TEMPLATES_DIR / "input_template.xlsx"
    wb.save(out)
    print(f"Template written → {out}")


if __name__ == "__main__":
    main()
